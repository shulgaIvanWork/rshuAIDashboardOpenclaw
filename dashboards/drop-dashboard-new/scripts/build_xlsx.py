"""
build_xlsx.py — генерирует Excel-отчёт из agg.json.
Требует: pip install openpyxl
Результат сохраняется в OUTPUT_DIR/XLSX_FILE (из config.py).
"""
import json, os
import config
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference, PieChart
from openpyxl.formatting.rule import ColorScaleRule

os.makedirs(config.OUTPUT_DIR, exist_ok=True)
agg = json.load(open(config.AGG_JSON, encoding="utf-8"))

NAVY = "1F2A44"; GOLD = "C8A45C"; GRAY = "F2F2F2"; WHITE = "FFFFFF"; KOM_C = "9B2D3C"
thin  = Side(style="thin", color="DDDDDD")
B     = Border(left=thin, right=thin, top=thin, bottom=thin)
HF    = Font(size=11, bold=True, color=WHITE)
HFILL     = PatternFill("solid", fgColor=NAVY)
HFILL_KOM = PatternFill("solid", fgColor=KOM_C)
TITLE     = Font(size=18, bold=True, color=NAVY)
SUB       = Font(size=11, italic=True, color="555555")

wb    = Workbook()
ytd   = agg["ytd"]; pw = agg["prev"]; cw = agg["cur"]; weeks = agg["weeks"]
kom_ytd  = agg["kom_ytd"]; kom_prev = agg["kom_prev"]; kom_cur = agg["kom_cur"]
prev_w   = agg["prev_week"]; cur_w = agg["cur_week"]

# ============ ЛИСТ 1: СВОДКА ============
ws = wb.active; ws.title = "Сводка"
ws.merge_cells("A1:F1")
ws["A1"] = f"Отчёт по продажам — {config.YEAR} год"; ws["A1"].font = TITLE
ws.merge_cells("A2:F2")
ws["A2"] = f"Битрикс24 · актуально на {agg['today']} · только сделки ≥ 1 ₽ · исключены «Pre Sale» и «КОМ (Sale)»"
ws["A2"].font = SUB

kpis = [
    ("Поступления YTD",   f"{ytd['postupleniya']:,.0f} ₽"),
    ("Релевантных WON",   f"{ytd['won_relevant_cnt']:,}"),
    ("Средний чек",       f"{ytd['avg_check']:,.0f} ₽"),
    ("Срок WON, дн.",     f"{ytd['avg_close_days_won']:.1f}"),
    ("Лидов YTD",         f"{agg['leads_ytd']:,}"),
    ("Конв. WON/(W+L)",   f"{ytd['conv_deal_pct']:.1f} %"),
    ("Сделок созд.",      f"{ytd['created_cnt']:,}"),
    ("Максимальный чек",  f"{ytd['max_check']:,.0f} ₽"),
]
for i, (k, v) in enumerate(kpis):
    col = 1 + (i % 4) * 2; row = 4 + (i // 4) * 3
    ws.cell(row=row,   column=col, value=k).font = Font(bold=True, color="555555", size=10)
    ws.cell(row=row+1, column=col, value=v).font = Font(bold=True, size=16, color=NAVY)
    ws.cell(row=row,   column=col).fill = PatternFill("solid", fgColor=GRAY)
    ws.cell(row=row+1, column=col).fill = PatternFill("solid", fgColor=GRAY)
    ws.merge_cells(start_row=row,   end_row=row,   start_column=col, end_column=col+1)
    ws.merge_cells(start_row=row+1, end_row=row+1, start_column=col, end_column=col+1)

start_row = 12
headers = ["Неделя","Период","MQL\n(новые в Pre Sale)","SQL\n(переданы в ОП)","Оплат\n(WON ≥1₽)",
           "Поступления, ₽","Сред.чек, ₽","Срок WON, дн.","MQL→SQL, %","SQL→Опл., %","Лидов всего","Конв.лид→опл., %"]
for j, h in enumerate(headers):
    c = ws.cell(row=start_row, column=j+1, value=h)
    c.font = HF; c.fill = HFILL; c.border = B
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[start_row].height = 40

for i, w in enumerate(weeks):
    r    = start_row + 1 + i
    cvL  = w["oplata"] / w["leads"] * 100 if w["leads"] else 0
    vals = [f"W{w['week']:02d}", w["label_dates"],
            w["mql"], w["sql"], w["oplata"],
            w["postupleniya"], w["avg_check"], w["avg_dur"],
            w["conv_mql_sql"], w["conv_sql_oplata"], w["leads"], cvL]
    for j, v in enumerate(vals):
        c = ws.cell(row=r, column=j+1, value=v); c.border = B
        if j == 5: c.number_format = '#,##0 ₽'
        if j == 6: c.number_format = '#,##0 ₽'
        if j in (2, 3, 4, 10): c.number_format = '#,##0'
        if j == 7: c.number_format = '0.0'
        if j in (8, 9, 11): c.number_format = '0.0 "%"'

r = start_row + 1 + len(weeks)
total_mql  = sum(w["mql"]          for w in weeks)
total_sql  = sum(w["sql"]          for w in weeks)
total_opl  = sum(w["oplata"]       for w in weeks)
total_pos  = sum(w["postupleniya"] for w in weeks)
total_leads= sum(w["leads"]        for w in weeks)
totals = ["ИТОГО","", total_mql, total_sql, total_opl, total_pos,
          total_pos/total_opl if total_opl else 0, ytd["avg_close_days_won"],
          total_sql/total_mql*100 if total_mql else 0,
          total_opl/total_sql*100 if total_sql else 0,
          total_leads, total_opl/total_leads*100 if total_leads else 0]
for j, v in enumerate(totals):
    c = ws.cell(row=r, column=j+1, value=v); c.border = B
    c.fill = PatternFill("solid", fgColor=GOLD); c.font = Font(bold=True, color=WHITE)
    if j in (5, 6): c.number_format = '#,##0 ₽'
    if j in (2, 3, 4, 10): c.number_format = '#,##0'
    if j == 7: c.number_format = '0.0'
    if j in (8, 9, 11): c.number_format = '0.0 "%"'

ws.conditional_formatting.add(f"F{start_row+1}:F{start_row+len(weeks)}",
    ColorScaleRule(start_type="min", start_color="FFFFFF",
                   mid_type="percentile", mid_value=50, mid_color="FFD180",
                   end_type="max", end_color="2E7D32"))
for i, w in enumerate([7,16,8,8,8,17,15,10,11,11,10,15]):
    ws.column_dimensions[get_column_letter(i+1)].width = w

# ============ ЛИСТ 2: КОМ ============
ws_k = wb.create_sheet("КОМ — отдельно")
ws_k.merge_cells("A1:D1"); ws_k["A1"] = "Воронка КОМ (Sale) — отдельный блок"
ws_k["A1"].font = Font(size=18, bold=True, color=KOM_C)
ws_k.merge_cells("A2:D2"); ws_k["A2"] = "«Копии для статистики» — КП и крупные корпоративные договоры"
ws_k["A2"].font = SUB

hdr = ["Показатель","YTD",
       f"W{prev_w:02d} ({weeks[-2]['label_dates'] if len(weeks)>=2 else ''})",
       f"W{cur_w:02d} ({weeks[-1]['label_dates']})"]
for j, h in enumerate(hdr):
    c = ws_k.cell(row=4, column=j+1, value=h); c.font = HF; c.fill = HFILL_KOM; c.border = B
    c.alignment = Alignment(horizontal="center", wrap_text=True)
ws_k.row_dimensions[4].height = 32

def add(ws_t, label, vals, fmt='#,##0', section=False):
    r = ws_t.max_row + 1
    for j, v in enumerate([label] + list(vals)):
        c = ws_t.cell(row=r, column=j+1, value=v); c.border = B
        if section:
            c.font = Font(bold=True, size=12, color=KOM_C)
        elif j == 0:
            c.font = Font(bold=True, color=NAVY); c.fill = PatternFill("solid", fgColor=GRAY)
        else:
            if fmt: c.number_format = fmt
            c.alignment = Alignment(horizontal="right")

ws_k.cell(row=ws_k.max_row+2, column=1, value="ПОСТУПЛЕНИЯ КОМ").font = Font(bold=True, size=12, color=KOM_C)
add(ws_k,"Поступления, ₽",     [kom_ytd["postupleniya"],    kom_prev["postupleniya"],    kom_cur["postupleniya"]],   '#,##0 ₽')
add(ws_k,"WON сделок",         [kom_ytd["won_relevant_cnt"],kom_prev["won_relevant_cnt"],kom_cur["won_relevant_cnt"]],'#,##0')
ws_k.cell(row=ws_k.max_row+2, column=1, value="СТОИМОСТЬ СДЕЛКИ КОМ").font = Font(bold=True, size=12, color=KOM_C)
add(ws_k,"Средний чек, ₽",     [kom_ytd["avg_check"],       kom_prev["avg_check"],       kom_cur["avg_check"]],      '#,##0 ₽')
add(ws_k,"Медианный чек, ₽",   [kom_ytd["median_check"],    kom_prev["median_check"],    kom_cur["median_check"]],   '#,##0 ₽')
add(ws_k,"Максимальный чек, ₽",[kom_ytd["max_check"],       kom_prev["max_check"],       kom_cur["max_check"]],      '#,##0 ₽')
ws_k.cell(row=ws_k.max_row+2, column=1, value="СКОРОСТЬ ЗАКРЫТИЯ КОМ").font = Font(bold=True, size=12, color=KOM_C)
add(ws_k,"Сред. время WON, дн.",[kom_ytd["avg_close_days_won"], kom_prev["avg_close_days_won"], kom_cur["avg_close_days_won"]], '0.0')
add(ws_k,"Медиана WON, дн.",    [kom_ytd["median_close_days_won"],kom_prev["median_close_days_won"],kom_cur["median_close_days_won"]],'0.0')
add(ws_k,"Сред. время LOSE, дн.",[kom_ytd["avg_close_days_lose"],kom_prev["avg_close_days_lose"],kom_cur["avg_close_days_lose"]],'0.0')
ws_k.cell(row=ws_k.max_row+2, column=1, value="КОНВЕРСИЯ КОМ").font = Font(bold=True, size=12, color=KOM_C)
add(ws_k,"Проигранных",        [kom_ytd["lose_cnt"],         kom_prev["lose_cnt"],         kom_cur["lose_cnt"]],        '#,##0')
add(ws_k,"Конв. WON/(WON+LOSE), %",[kom_ytd["conv_deal_pct"],kom_prev["conv_deal_pct"],kom_cur["conv_deal_pct"]],'0.0 "%"')

for col, w in [("A",38),("B",22),("C",24),("D",24)]:
    ws_k.column_dimensions[col].width = w

# ============ ЛИСТ 3: ВОРОНКА MQL→SQL ============
ws_f = wb.create_sheet("Воронка MQL-SQL-Оплата")
ws_f.merge_cells("A1:G1"); ws_f["A1"] = "Воронка MQL → SQL → Оплата по неделям"; ws_f["A1"].font = TITLE
ws_f.merge_cells("A2:G2"); ws_f["A2"] = "MQL = новые в Pre Sale. SQL = WON в Pre Sale. Оплата = WON ≥1₽."
ws_f["A2"].font = SUB; ws_f["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws_f.row_dimensions[2].height = 35

for j, h in enumerate(["Неделя","Период","MQL","SQL","Оплата","MQL→SQL, %","SQL→Опл., %","Скорость PreSale, дн."]):
    c = ws_f.cell(row=4, column=j+1, value=h); c.font = HF; c.fill = HFILL; c.border = B
    c.alignment = Alignment(horizontal="center", wrap_text=True)
ws_f.row_dimensions[4].height = 32
for i, w in enumerate(weeks):
    r    = 5 + i
    vals = [f"W{w['week']:02d}", w["label_dates"], w["mql"], w["sql"], w["oplata"],
            w["conv_mql_sql"], w["conv_sql_oplata"], w["avg_presale_dur"]]
    for j, v in enumerate(vals):
        c = ws_f.cell(row=r, column=j+1, value=v); c.border = B
        if j in (2, 3, 4): c.number_format = '#,##0'
        if j in (5, 6):    c.number_format = '0.0 "%"'
        if j == 7:          c.number_format = '0.0'
for i, w in enumerate([7,16,8,8,8,12,12,18]):
    ws_f.column_dimensions[get_column_letter(i+1)].width = w

ch = BarChart(); ch.title = "Воронка MQL → SQL → Оплата"
ch.type = "col"; ch.style = 12; ch.height = 11; ch.width = 24
for col in range(3, 6):
    d = Reference(ws_f, min_col=col, min_row=4, max_row=4+len(weeks))
    ch.add_data(d, titles_from_data=True)
ch.set_categories(Reference(ws_f, min_col=1, min_row=5, max_row=4+len(weeks)))
ws_f.add_chart(ch, "J4")

# ============ ЛИСТ 4: ИСТОЧНИКИ ============
ws_s = wb.create_sheet("Источники")
ws_s.merge_cells("A1:E1"); ws_s["A1"] = "Рейтинг источников по поступлениям (YTD)"; ws_s["A1"].font = TITLE
ws_s.merge_cells("A2:E2"); ws_s["A2"] = "Источник = SOURCE_ID сделки. «Лидов» — кол-во лидов с этим источником."
ws_s["A2"].font = SUB
for j, h in enumerate(["#","Источник","Сделок WON","Лидов","Поступления, ₽","Доля, %"]):
    c = ws_s.cell(row=4, column=j+1, value=h); c.font = HF; c.fill = HFILL; c.border = B
    c.alignment = Alignment(horizontal="center")
total_src = sum(t["deals"] for t in agg["src_rating"]) or 1
for i, t in enumerate(agg["src_rating"][:30]):
    r = 5 + i
    ws_s.cell(row=r, column=1, value=i+1).border = B
    ws_s.cell(row=r, column=2, value=t.get("name","—")).border = B
    ws_s.cell(row=r, column=3, value=t.get("deals",0)).border = B
    ws_s.cell(row=r, column=4, value=t.get("leads",0)).border = B
    c = ws_s.cell(row=r, column=5, value=t.get("postupleniya",0)); c.border = B; c.number_format = '#,##0 ₽'
    c = ws_s.cell(row=r, column=6, value=t.get("postupleniya",0)/total_src*100); c.border = B; c.number_format = '0.0 "%"'
for col, w in [("A",5),("B",45),("C",12),("D",10),("E",18),("F",10)]:
    ws_s.column_dimensions[col].width = w

# ============ ЛИСТ 5: B2B vs B2C ============
ws_b = wb.create_sheet("B2B vs B2C")
ws_b.merge_cells("A1:D1"); ws_b["A1"] = "Разбивка поступлений B2B vs B2C"; ws_b["A1"].font = TITLE
ws_b.merge_cells("A2:D2"); ws_b["A2"] = "B2B = компания (COMPANY_ID) или сумма ≥ 50 000 ₽ или воронка КОМ. B2C = всё остальное."
ws_b["A2"].font = SUB
for j, h in enumerate(["Период","Тип","Сделок","Поступления, ₽","Доля, %"]):
    c = ws_b.cell(row=4, column=j+1, value=h); c.font = HF; c.fill = HFILL; c.border = B
    c.alignment = Alignment(horizontal="center")
r = 5
for d, period in [(agg["btype_ytd"],"YTD"),(agg["btype_prev"],f"W{prev_w:02d}")]:
    b2b = d.get("B2B",{"cnt":0,"sum":0}); b2c = d.get("B2C",{"cnt":0,"sum":0})
    tot = b2b["sum"] + b2c["sum"] or 1
    for t, v in [("B2B",b2b),("B2C",b2c)]:
        ws_b.cell(row=r,column=1,value=period).border=B
        c=ws_b.cell(row=r,column=2,value=t); c.border=B
        c.fill=PatternFill("solid",fgColor="E3F2FD" if t=="B2B" else "FFF3E0")
        ws_b.cell(row=r,column=3,value=v["cnt"]).border=B
        c=ws_b.cell(row=r,column=4,value=v["sum"]); c.border=B; c.number_format='#,##0 ₽'
        c=ws_b.cell(row=r,column=5,value=v["sum"]/tot*100); c.border=B; c.number_format='0.0 "%"'
        r+=1
    r+=1
for col, w in [("A",12),("B",10),("C",10),("D",20),("E",10)]:
    ws_b.column_dimensions[col].width = w

# ============ ЛИСТ 6: ФОРМАТЫ ============
ws_fm = wb.create_sheet("Форматы")
ws_fm.merge_cells("A1:D1"); ws_fm["A1"] = "Разбивка по форматам обучения"; ws_fm["A1"].font = TITLE
ws_fm.merge_cells("A2:D2"); ws_fm["A2"] = "ОМ (Онлайн) + ООМ (Очное) + СДО + КОМ (корпоративные договоры)"
ws_fm["A2"].font = SUB
for j, h in enumerate(["Период","Формат","Сделок","Поступления, ₽","Доля, %"]):
    c = ws_fm.cell(row=4, column=j+1, value=h); c.font = HF; c.fill = HFILL; c.border = B
    c.alignment = Alignment(horizontal="center")
r = 5
FCOLORS = {"ОМ (Онлайн)":"E8F5E9","ООМ (Очное)":"E3F2FD","СДО":"FFF3E0","КОМ":"FFEBEE"}
for src, period in [(agg["fmt_ytd"],"YTD"),(agg["fmt_prev"],f"W{prev_w:02d}")]:
    items = sorted([(k,v) for k,v in src.items() if k!="period"], key=lambda x:-x[1]["sum"])
    tot   = sum(v["sum"] for _, v in items) or 1
    for k, v in items:
        ws_fm.cell(row=r,column=1,value=period).border=B
        c=ws_fm.cell(row=r,column=2,value=k); c.border=B
        c.fill=PatternFill("solid",fgColor=FCOLORS.get(k,"F5F5F5"))
        ws_fm.cell(row=r,column=3,value=v["cnt"]).border=B
        c=ws_fm.cell(row=r,column=4,value=v["sum"]); c.border=B; c.number_format='#,##0 ₽'
        c=ws_fm.cell(row=r,column=5,value=v["sum"]/tot*100); c.border=B; c.number_format='0.0 "%"'
        r+=1
    r+=1
for col, w in [("A",12),("B",15),("C",10),("D",20),("E",10)]:
    ws_fm.column_dimensions[col].width = w

# ============ ЛИСТ 7: ТОП-20 ПРОДУКТОВ ============
ws_p = wb.create_sheet("ТОП-20 продуктов")
ws_p.merge_cells("A1:F1"); ws_p["A1"] = "ТОП-20 продуктов по поступлениям (YTD, без КОМ и Post Sale)"; ws_p["A1"].font = TITLE
ws_p.merge_cells("A2:F2"); ws_p["A2"] = "Названия нормализованы — даты/города/договоры удалены."
ws_p["A2"].font = SUB
for j, h in enumerate(["#","Продукт","Сделок","Поступления, ₽","Средний чек, ₽","Доля, %"]):
    c = ws_p.cell(row=4, column=j+1, value=h); c.font = HF; c.fill = HFILL; c.border = B
total_top = sum(t["deals"] for t in agg["top_products"]) or 1
for i, tp in enumerate(agg["top_products"]):
    r = 5 + i
    ws_p.cell(row=r,column=1,value=i+1).border=B
    ws_p.cell(row=r,column=2,value=tp.get("name","—")).border=B
    ws_p.cell(row=r,column=3,value=tp.get("deals",0)).border=B
    c=ws_p.cell(row=r,column=4,value=tp.get("sum",0)); c.border=B; c.number_format='#,##0 ₽'
    c=ws_p.cell(row=r,column=5,value=tp.get("avg_check",0)); c.border=B; c.number_format='#,##0 ₽'
    c=ws_p.cell(row=r,column=6,value=tp.get("sum",0)/total_top*100); c.border=B; c.number_format='0.0"%"'
for col, w in [("A",5),("B",65),("C",9),("D",18),("E",18),("F",10)]:
    ws_p.column_dimensions[col].width = w

# ============ ЛИСТ 8: МЕНЕДЖЕРЫ ============
ws_m = wb.create_sheet("Менеджеры")
ws_m.merge_cells("A1:H1"); ws_m["A1"] = "Поступления по менеджерам YTD (основная воронка)"; ws_m["A1"].font = TITLE
ws_m.merge_cells("A2:H2"); ws_m["A2"] = "Лиды — crm.lead.list по ASSIGNED_BY_ID. Конверсия = WON/(WON+Проигр)."
ws_m["A2"].font = SUB
for j, h in enumerate(["#","Менеджер","Лидов CRM","WON","Проигр","Конв. %","Ср. чек, ₽","Поступления, ₽"]):
    c = ws_m.cell(row=4, column=j+1, value=h); c.font = HF; c.fill = HFILL; c.border = B
for i, m in enumerate(agg["mgr_top"]):
    if isinstance(m, dict):
        vals = [i+1, m.get("name","—"), m.get("leads",0), m.get("won",0),
                m.get("lost",0), m.get("conv_pct",0), m.get("avg_check",0), m.get("postupleniya",0)]
    else:
        vals = [i+1, m[0], "—", m[1], "—", "—", int(m[2]/m[1]) if m[1] else 0, m[2]]
    r = 5 + i
    for j, v in enumerate(vals):
        c = ws_m.cell(row=r, column=j+1, value=v); c.border = B
        if j == 5 and isinstance(v,(int,float)): c.number_format = '0.0"%"'
        if j in (6,7) and isinstance(v,(int,float)): c.number_format = '#,##0 ₽'
ws_m.conditional_formatting.add(f"H5:H{4+len(agg['mgr_top'])}",
    ColorScaleRule(start_type="min",start_color="FFFFFF",
                   mid_type="percentile",mid_value=50,mid_color="C8E6C9",
                   end_type="max",end_color="1B5E20"))
for col, w in [("A",5),("B",35),("C",13),("D",8),("E",10),("F",10),("G",18),("H",20)]:
    ws_m.column_dimensions[col].width = w

out_path = os.path.join(config.OUTPUT_DIR, config.XLSX_FILE)
wb.save(out_path)
print(f"Excel готов → {out_path}")
